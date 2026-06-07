"""Edge cases for the Excel batch export sheet-name collision logic.

The batch xlsx generator has two failure modes that the rest of the
suite does not exercise:

1. A user-named job called exactly ``"Combined"`` (the name the
   flatten branch also uses) — without defensive de-duplication the
   explicit ``create_sheet(title=...)`` call would raise
   ``InvalidWorksheetTitle``.
2. A user-named job called exactly ``"Sheet"`` — the per-job branch
   reserves the name ``"Sheet"`` as the empty-name fallback and must
   suffix it.
3. A 31-character job name — Excel's hard cap on sheet-name length
   is 31 characters; the per-job branch truncates with ``[:31]`` but
   must still avoid collisions among siblings.
4. A whitespace-only job name — the ``or "Sheet"`` fallback should
   kick in only when the name is empty/falsy, not when it is
   whitespace (the implementation does not strip, so a whitespace
   name passes through as the base and gets suffixed). Either way
   the result must be a valid xlsx archive.

Each test asserts (a) the endpoint returns 200, (b) the response is
a valid openpyxl-readable xlsx (i.e. ``content[:2] == b"PK"``), and
(c) the workbook has the expected number of sheets so a future
regression that silently swallows sheet creation is caught.
"""

from __future__ import annotations

import io
from typing import Any

import pytest
import pytest_asyncio
from app.models import Job, JobStatus
from app.routers.exports import create_exports_router
from fastapi import FastAPI
from openpyxl import load_workbook
from pydantic import BaseModel


def _make_job(
    job_id: str,
    name: str = "test-job",
    results: list[dict[str, Any]] | None = None,
) -> Job:
    return Job(
        id=job_id,
        name=name,
        status=JobStatus.COMPLETED,
        results=results or [{"x": "1"}],
        schema_fields=[],
        urls=["https://example.com"],
        results_on_disk=False,
    )


class _BatchClient(BaseModel):
    """Minimal httpx-free wrapper that mimics the AsyncClient .post shape."""

    app: Any
    transport: Any
    client: Any

    model_config = {"arbitrary_types_allowed": True}

    async def post(self, path: str, json: dict) -> Any:
        return await self.client.post(path, json=json)


@pytest_asyncio.fixture
async def build_client():
    """Return a factory that builds an httpx AsyncClient over a fresh app."""

    def _factory(jobs_store: dict[str, Job]) -> _BatchClient:
        from httpx import ASGITransport, AsyncClient

        router = create_exports_router(jobs_store)
        test_app = FastAPI()
        test_app.include_router(router)
        transport = ASGITransport(app=test_app)
        client = AsyncClient(transport=transport, base_url="http://testserver")
        return _BatchClient(app=test_app, transport=transport, client=client)

    return _factory


def _assert_valid_xlsx(content: bytes, expected_sheet_count: int) -> None:
    """Open the bytes as an xlsx and assert the sheet count."""
    assert content[:2] == b"PK", "Response is not an xlsx archive (missing PK header)"
    wb = load_workbook(io.BytesIO(content), read_only=True)
    try:
        # Some openpyxl versions expose ``sheetnames`` even on
        # read-only / write-only workbooks, but ``_sheets`` is the
        # lower-level fallback. Use whichever is populated.
        names = wb.sheetnames if getattr(wb, "sheetnames", None) else [ws.title for ws in wb.worksheets]
        assert len(names) == expected_sheet_count, f"Expected {expected_sheet_count} sheet(s), got {len(names)}: {names!r}"
        # Each sheet title must be a non-empty string and respect
        # the 31-character Excel cap.
        for title in names:
            assert isinstance(title, str)
            assert 1 <= len(title) <= 31, f"Invalid sheet title: {title!r}"
    finally:
        wb.close()


# ─── Job named "Combined" with flatten=True ────────────────────────────


@pytest.mark.asyncio
async def test_batch_xlsx_combined_named_job_with_flatten(build_client) -> None:
    """A user job named ``"Combined"`` must not collide with the flatten sheet.

    The flatten branch hard-codes ``base = "Combined"``. If a
    user-named job is also called ``"Combined"`` the explicit
    ``create_sheet(title="Combined")`` would fail the second time
    (or in the per-job branch after flatten resolves). The
    pre-registration in ``used_flatten_names`` is what makes this
    safe.
    """
    jobs_store: dict[str, Job] = {
        "user-combined": _make_job("user-combined", name="Combined"),
        "other": _make_job("other", name="Other"),
    }
    bc = build_client(jobs_store)
    try:
        resp = await bc.post(
            "/api/exports/batch",
            json={"job_ids": ["user-combined", "other"], "format": "xlsx", "flatten": True},
        )
        assert resp.status_code == 200
        # 1 sheet (the flatten target) — not 2, because flatten=True.
        _assert_valid_xlsx(resp.content, expected_sheet_count=1)
    finally:
        await bc.client.aclose()


# ─── Job named "Sheet" ────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_batch_xlsx_job_named_sheet(build_client) -> None:
    """A user job named ``"Sheet"`` must not collide with the per-job fallback.

    The per-job branch treats ``"Sheet"`` as a reserved name (the
    fallback for ``job_name or "Sheet"``) and disambiguates with a
    numeric suffix. The output must therefore be a single, valid
    sheet (with a suffix appended) — not a 500 from
    ``InvalidWorksheetTitle``.
    """
    jobs_store: dict[str, Job] = {
        "sheet": _make_job("sheet", name="Sheet"),
    }
    bc = build_client(jobs_store)
    try:
        resp = await bc.post(
            "/api/exports/batch",
            json={"job_ids": ["sheet"], "format": "xlsx", "flatten": False},
        )
        assert resp.status_code == 200
        # 1 sheet: the per-job branch renames "Sheet" to e.g. "Sheet (2)".
        _assert_valid_xlsx(resp.content, expected_sheet_count=1)
        wb = load_workbook(io.BytesIO(resp.content), read_only=True)
        try:
            assert wb.sheetnames[0] != "Sheet", f"Sheet name should be disambiguated, got {wb.sheetnames[0]!r}"
        finally:
            wb.close()
    finally:
        await bc.client.aclose()


# ─── 31-character job name ────────────────────────────────────────────


@pytest.mark.asyncio
async def test_batch_xlsx_31_char_job_name(build_client) -> None:
    """A job name at the Excel sheet-name length cap must round-trip cleanly."""
    long_name = "A" * 31
    jobs_store: dict[str, Job] = {
        "long": _make_job("long", name=long_name),
    }
    bc = build_client(jobs_store)
    try:
        resp = await bc.post(
            "/api/exports/batch",
            json={"job_ids": ["long"], "format": "xlsx", "flatten": False},
        )
        assert resp.status_code == 200
        _assert_valid_xlsx(resp.content, expected_sheet_count=1)
    finally:
        await bc.client.aclose()


@pytest.mark.asyncio
async def test_batch_xlsx_31_char_names_collision_avoided(build_client) -> None:
    """Two jobs whose names share a 31-char prefix must not collide.

    The collision-avoidance loop appends a numeric suffix
    (e.g. ``" (2)"``) when a duplicate is detected. Without it,
    openpyxl raises ``InvalidWorksheetTitle`` on the second
    ``create_sheet`` call. With it, the workbook contains two
    distinct sheets.
    """
    long_name = "B" * 31
    jobs_store: dict[str, Job] = {
        "first": _make_job("first", name=long_name),
        "second": _make_job("second", name=long_name),
    }
    bc = build_client(jobs_store)
    try:
        resp = await bc.post(
            "/api/exports/batch",
            json={"job_ids": ["first", "second"], "format": "xlsx", "flatten": False},
        )
        assert resp.status_code == 200
        _assert_valid_xlsx(resp.content, expected_sheet_count=2)
        wb = load_workbook(io.BytesIO(resp.content), read_only=True)
        try:
            names = wb.sheetnames
            assert len(set(names)) == 2, f"Sheet names must be unique, got {names!r}"
        finally:
            wb.close()
    finally:
        await bc.client.aclose()


# ─── Whitespace-only job name ─────────────────────────────────────────


@pytest.mark.asyncio
async def test_batch_xlsx_whitespace_only_job_name(build_client) -> None:
    """A whitespace-only job name must yield a valid xlsx.

    The implementation does not strip whitespace, so a
    whitespace-only name (e.g. ``"   "``) is truthy and passes the
    ``or "Sheet"`` fallback. That means the whitespace string is
    used as the base sheet name. Excel accepts spaces in sheet
    names, so this must produce a 200 with a valid workbook — not
    a 500 from openpyxl validation.
    """
    jobs_store: dict[str, Job] = {
        "ws": _make_job("ws", name="   "),
    }
    bc = build_client(jobs_store)
    try:
        resp = await bc.post(
            "/api/exports/batch",
            json={"job_ids": ["ws"], "format": "xlsx", "flatten": False},
        )
        assert resp.status_code == 200
        _assert_valid_xlsx(resp.content, expected_sheet_count=1)
    finally:
        await bc.client.aclose()

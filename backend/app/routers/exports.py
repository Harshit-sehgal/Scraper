import csv
import datetime
import io
import json
import logging
from collections.abc import AsyncIterator
from typing import Annotated, Any

from fastapi import APIRouter, Depends, HTTPException, Response
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel, Field
from starlette.concurrency import run_in_threadpool

from app.config import settings
from app.utils.export import safe_export_filename
from app.utils.rbac import UserRole, require_role

logger = logging.getLogger(__name__)

_PAGINATION_CHUNK_SIZE = 500


def _user_fieldnames(results_list: list[dict]) -> list[str]:
    """Return field names from the first record, filtering out internal system fields (keys starting with ``_``).

    When no schema is defined, we fall back to the keys of the first result
    record.  Internal metadata fields (``_acquisition_lineage``,
    ``_provenance``, …) are stripped so they never leak into user-facing
    exports.
    """
    if not results_list:
        return []
    return [k for k in results_list[0] if not k.startswith("_")]


def _strip_system_fields(records: list[dict]) -> list[dict]:
    """Return a deep-ish copy of *records* with all keys starting with ``_`` removed."""
    return [{k: v for k, v in r.items() if not k.startswith("_")} for r in records]


def _flat_row(row: dict, fieldnames: list[str]) -> dict:
    """Flatten list values in a row to comma-separated strings and escape formula injection."""
    flat = {}
    for k in fieldnames:
        v = row.get(k)
        if isinstance(v, list):
            flat[k] = _safe_cell(", ".join(str(i) for i in v))
        else:
            flat[k] = _safe_cell(v)
    return flat


_DANGEROUS_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value):
    """Escape spreadsheet formula-injection prefixes so exported CSV / Excel files
    do not execute malicious formulas when opened in Excel, Sheets, or LibreOffice.

    If *value* is a string that starts with a dangerous prefix (``=``, ``+``, ``-``,
    ``@``, tab, or carriage return), prepend a single quote so the spreadsheet
    software treats it as plain text.

    Non-string values are returned unchanged.
    """
    if isinstance(value, str) and value.startswith(_DANGEROUS_PREFIXES):
        return "'" + value
    return value


class BatchExportRequest(BaseModel):
    """Request body for batch export endpoint."""

    job_ids: list[str] = Field(
        ...,
        min_length=1,
        max_length=50,
        description="Job IDs to include in the batch export",
    )
    format: str = Field(
        "csv",
        description="Export format: csv, json, or xlsx",
    )
    flatten: bool = Field(
        True,
        description="When True, all results are combined into a single output. "
        "When False, CSV uses separator rows and Excel uses one sheet per job.",
    )


_SOURCE_JOB_FIELD = "_source_job"
"""Field name injected into each result row to identify the source job."""


def _record_export_outcome(fmt: str, success: bool) -> None:
    """Record an export generation outcome for the metrics endpoint.

    Mirrors :func:`app.metrics_collector.record_export_outcome`. The
    function is intentionally tolerant: a broken metrics subsystem
    must never turn a successful export into a 5xx.
    """
    try:
        from app.metrics_collector import record_export_outcome

        record_export_outcome(fmt, success)
    except Exception:
        logger.debug("Failed to record export outcome metric for %s", fmt)


def create_exports_router(jobs_store: dict):
    router = APIRouter()

    # Single-process lock guarding refresh-from-repo writes into ``jobs_store``.
    # Without this, two concurrent export requests can race and observe a
    # partially-overwritten job (or worse, raise ``RuntimeError: dictionary
    # changed size during iteration`` when a reader iterates ``jobs_store``).
    import threading

    _store_lock = threading.Lock()

    def _refresh_job_for_export(job_id: str) -> None:
        """Refresh job from repository in worker mode to avoid stale exports.

        Uses the targeted ``get_job()`` read instead of loading all jobs
        and filtering client-side. This helper is synchronous; async
        export routes call it via ``run_in_threadpool``.
        """
        if settings.WORKER_QUEUE:
            try:
                from app.storage_interface import get_job_repository

                repo = get_job_repository()
                fresh = repo.get_job(job_id)
                if fresh is not None:
                    with _store_lock:
                        jobs_store[job_id] = fresh
            except Exception:
                logger.debug("Failed to refresh job %s from repo for export", job_id)

    @router.get("/api/jobs/{job_id}/export/csv")
    async def export_csv(
        job_id: str,
        _role: Annotated[UserRole, Depends(require_role([UserRole.ADMIN, UserRole.OPERATOR]))],
    ):
        try:
            result = await _export_csv_impl(job_id)
        except HTTPException:
            _record_export_outcome("csv", False)
            raise
        except Exception:
            _record_export_outcome("csv", False)
            raise
        else:
            _record_export_outcome("csv", True)
            return result

    async def _export_csv_impl(job_id: str):
        await run_in_threadpool(_refresh_job_for_export, job_id)
        job = jobs_store.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")

        if job.results_on_disk:
            from app.utils.job_results_store import (
                load_paginated_job_results_from_disk,
            )

            # Load the first page to determine headers and total count
            first_page, total = await run_in_threadpool(
                load_paginated_job_results_from_disk,
                job.id,
                limit=_PAGINATION_CHUNK_SIZE,
                offset=0,
                file_path=job.results_file_path,
            )
            if not first_page:
                raise HTTPException(status_code=400, detail="No results to export")

            fieldnames = [f.name for f in job.schema_fields] if job.schema_fields else _user_fieldnames(first_page)

            async def _stream_csv_from_disk() -> AsyncIterator[str]:
                output = io.StringIO()
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                yield output.getvalue()
                output.seek(0)
                output.truncate()

                # Yield first page
                for row in first_page:
                    writer.writerow(_flat_row(row, fieldnames))
                yield output.getvalue()
                output.seek(0)
                output.truncate()

                # Stream remaining pages
                offset = _PAGINATION_CHUNK_SIZE
                while offset < total:
                    page, _ = await run_in_threadpool(
                        load_paginated_job_results_from_disk,
                        job.id,
                        limit=_PAGINATION_CHUNK_SIZE,
                        offset=offset,
                        file_path=job.results_file_path,
                    )
                    if not page:
                        break
                    for row in page:
                        writer.writerow(_flat_row(row, fieldnames))
                    yield output.getvalue()
                    output.seek(0)
                    output.truncate()
                    offset += _PAGINATION_CHUNK_SIZE

            return StreamingResponse(
                _stream_csv_from_disk(),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="{safe_export_filename(job.name, "csv")}"'},
            )

        # In-memory results (small dataset)
        if not job.results:
            raise HTTPException(status_code=400, detail="No results to export")

        output = io.StringIO()
        fieldnames = [f.name for f in job.schema_fields] if job.schema_fields else _user_fieldnames(job.results)
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for row in job.results:
            writer.writerow(_flat_row(row, fieldnames))

        output.seek(0)
        return Response(
            content=output.getvalue(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{safe_export_filename(job.name, "csv")}"'},
        )

    @router.get("/api/jobs/{job_id}/export/json")
    async def export_json(
        job_id: str,
        _role: Annotated[UserRole, Depends(require_role([UserRole.ADMIN, UserRole.OPERATOR]))],
    ):
        try:
            result = await _export_json_impl(job_id)
        except HTTPException:
            _record_export_outcome("json", False)
            raise
        except Exception:
            _record_export_outcome("json", False)
            raise
        else:
            _record_export_outcome("json", True)
            return result

    async def _export_json_impl(job_id: str):
        await run_in_threadpool(_refresh_job_for_export, job_id)
        job = jobs_store.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")

        if job.results_on_disk:
            from app.utils.job_results_store import (
                load_paginated_job_results_from_disk,
            )

            first_page, total = await run_in_threadpool(
                load_paginated_job_results_from_disk,
                job.id,
                limit=_PAGINATION_CHUNK_SIZE,
                offset=0,
                file_path=job.results_file_path,
            )
            if not first_page:
                raise HTTPException(status_code=400, detail="No results to export")

            async def _stream_json_from_disk() -> AsyncIterator[str]:
                yield "[\n"
                # Yield first page
                for i, row in enumerate(first_page):
                    cleaned = _strip_system_fields([row])[0]
                    prefix = "  " if i == 0 else ",  "
                    yield prefix + json.dumps(cleaned, indent=2).replace("\n", "\n  ") + "\n"

                # Stream remaining pages
                offset = _PAGINATION_CHUNK_SIZE
                idx = len(first_page)
                while offset < total:
                    page, _ = await run_in_threadpool(
                        load_paginated_job_results_from_disk,
                        job.id,
                        limit=_PAGINATION_CHUNK_SIZE,
                        offset=offset,
                        file_path=job.results_file_path,
                    )
                    if not page:
                        break
                    for row in page:
                        cleaned = _strip_system_fields([row])[0]
                        prefix = ",  " if idx > 0 else "  "
                        yield prefix + json.dumps(cleaned, indent=2).replace("\n", "\n  ") + "\n"
                        idx += 1
                    offset += _PAGINATION_CHUNK_SIZE
                yield "]\n"

            return StreamingResponse(
                _stream_json_from_disk(),
                media_type="application/json",
                headers={"Content-Disposition": f'attachment; filename="{safe_export_filename(job.name, "json")}"'},
            )

        # In-memory results
        if not job.results:
            raise HTTPException(status_code=400, detail="No results to export")

        cleaned = _strip_system_fields(list(job.results))
        json_content = json.dumps(cleaned, indent=2)
        return Response(
            content=json_content,
            media_type="application/json",
            headers={"Content-Disposition": f'attachment; filename="{safe_export_filename(job.name, "json")}"'},
        )

    @router.get("/api/jobs/{job_id}/export/excel")
    async def export_excel(
        job_id: str,
        _role: Annotated[UserRole, Depends(require_role([UserRole.ADMIN, UserRole.OPERATOR]))],
    ):
        try:
            result = await _export_excel_impl(job_id)
        except HTTPException:
            _record_export_outcome("excel", False)
            raise
        except Exception:
            _record_export_outcome("excel", False)
            raise
        _record_export_outcome("excel", True)
        return result

    async def _export_excel_impl(job_id: str):
        await run_in_threadpool(_refresh_job_for_export, job_id)
        job = jobs_store.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job not found")

        def _build_excel_content():
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title="Scraped Data")

            if job.results_on_disk:
                from app.utils.job_results_store import load_paginated_job_results_from_disk

                # Load the first page to determine headers and total count
                first_page, total = load_paginated_job_results_from_disk(
                    job.id,
                    limit=_PAGINATION_CHUNK_SIZE,
                    offset=0,
                    file_path=job.results_file_path,
                )
                if not first_page:
                    return None  # validated above or empty

                fieldnames = [f.name for f in job.schema_fields] if job.schema_fields else _user_fieldnames(first_page)

                # Write headers
                ws.append(fieldnames)

                # Write first page data
                for row in first_page:
                    row_values = []
                    for field in fieldnames:
                        value = row.get(field)
                        if isinstance(value, list):
                            value = _safe_cell(", ".join(str(i) for i in value if i is not None))
                        else:
                            value = _safe_cell(value)
                        row_values.append(value)
                    ws.append(row_values)

                # Stream remaining pages
                offset = _PAGINATION_CHUNK_SIZE
                while offset < total:
                    page, _ = load_paginated_job_results_from_disk(
                        job.id,
                        limit=_PAGINATION_CHUNK_SIZE,
                        offset=offset,
                        file_path=job.results_file_path,
                    )
                    if not page:
                        break
                    for row in page:
                        row_values = []
                        for field in fieldnames:
                            value = row.get(field)
                            if isinstance(value, list):
                                value = _safe_cell(", ".join(str(i) for i in value if i is not None))
                            else:
                                value = _safe_cell(value)
                            row_values.append(value)
                        ws.append(row_values)
                    offset += _PAGINATION_CHUNK_SIZE
            else:
                # In-memory results
                results_list = list(job.results)
                if not results_list:
                    return None  # no results to export

                fieldnames = [f.name for f in job.schema_fields] if job.schema_fields else _user_fieldnames(results_list)

                # Write headers
                ws.append(fieldnames)

                # Write data
                for row in results_list:
                    row_values = []
                    for field in fieldnames:
                        value = row.get(field)
                        if isinstance(value, list):
                            value = _safe_cell(", ".join(str(i) for i in value if i is not None))
                        else:
                            value = _safe_cell(value)
                        row_values.append(value)
                    ws.append(row_values)

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        # Validate data exists before entering threadpool
        if job.results_on_disk:
            from app.utils.job_results_store import load_paginated_job_results_from_disk

            _first_page, _ = await run_in_threadpool(
                load_paginated_job_results_from_disk,
                job.id,
                limit=1,
                offset=0,
                file_path=job.results_file_path,
            )
            if not _first_page:
                raise HTTPException(status_code=400, detail="No results to export")
        elif not job.results:
            raise HTTPException(status_code=400, detail="No results to export")

        content_bytes = await run_in_threadpool(_build_excel_content)

        return Response(
            content=content_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_export_filename(job.name, "xlsx")}"'},
        )

    # ─── Batch Export ────────────────────────────────────────────────────

    @router.post("/api/exports/batch")
    async def batch_export(
        body: BatchExportRequest,
        _role: Annotated[UserRole, Depends(require_role([UserRole.ADMIN, UserRole.OPERATOR]))],
    ):
        try:
            result = await _batch_export_impl(body)
        except HTTPException:
            _record_export_outcome(f"batch_{body.format}", False)
            raise
        except Exception:
            _record_export_outcome(f"batch_{body.format}", False)
            raise
        else:
            _record_export_outcome(f"batch_{body.format}", True)
            return result

    async def _batch_export_impl(body: BatchExportRequest):
        fmt = body.format.lower()
        if fmt not in ("csv", "json", "xlsx"):
            raise HTTPException(status_code=400, detail=f"Unsupported format '{fmt}'. Supported: csv, json, xlsx")

        # Resolve all jobs — fail fast on any missing ID.
        missing: list[str] = []
        for _jid in body.job_ids:
            if _jid not in jobs_store:
                missing.append(_jid)
            else:
                await run_in_threadpool(_refresh_job_for_export, _jid)

        if missing:
            raise HTTPException(
                status_code=404,
                detail=f"Jobs not found: {', '.join(missing)}",
            )

        # Resolve job metadata (name, on_disk, file_path) without loading results.
        job_meta: list[tuple[str, str, bool, str | None]] = []
        for _jid in body.job_ids:
            job = jobs_store.get(_jid)
            if not job:
                job_meta.append((_jid, _jid, False, None))
            else:
                job_meta.append((_jid, job.name or _jid, bool(job.results_on_disk), job.results_file_path))

        # ── Streaming fieldnames discovery ──────────────────────────────
        # Load only the first page of each job to discover the union of
        # fieldnames.  This avoids loading all results into memory.
        fieldnames: list[str] = []
        seen: set[str] = set()
        has_any_data = False
        for jid, _jname, on_disk, fpath in job_meta:
            if on_disk:
                from app.utils.job_results_store import load_paginated_job_results_from_disk

                first_page, _ = await run_in_threadpool(
                    load_paginated_job_results_from_disk,
                    jid,
                    limit=_PAGINATION_CHUNK_SIZE,
                    offset=0,
                    file_path=fpath,
                )
                if first_page:
                    has_any_data = True
                    cleaned = _strip_system_fields(first_page)
                    for row in cleaned:
                        for k in row:
                            if k not in seen:
                                seen.add(k)
                                fieldnames.append(k)
            else:
                job = jobs_store.get(jid)
                if job and job.results:
                    has_any_data = True
                    sample = list(job.results)[:_PAGINATION_CHUNK_SIZE]
                    cleaned = _strip_system_fields(sample)
                    for row in cleaned:
                        for k in row:
                            if k not in seen:
                                seen.add(k)
                                fieldnames.append(k)

        if not has_any_data:
            raise HTTPException(status_code=400, detail="None of the specified jobs have results to export")

        # Also discover fieldnames from subsequent pages (small overhead).
        for jid, _jname, on_disk, fpath in job_meta:
            if on_disk:
                from app.utils.job_results_store import load_paginated_job_results_from_disk

                offset = _PAGINATION_CHUNK_SIZE
                while True:
                    page, total = await run_in_threadpool(
                        load_paginated_job_results_from_disk,
                        jid,
                        limit=_PAGINATION_CHUNK_SIZE,
                        offset=offset,
                        file_path=fpath,
                    )
                    if not page:
                        break
                    cleaned = _strip_system_fields(page)
                    for row in cleaned:
                        for k in row:
                            if k not in seen:
                                seen.add(k)
                                fieldnames.append(k)
                    offset += _PAGINATION_CHUNK_SIZE
                    if offset >= total:
                        break

        # ── Streaming page generator ────────────────────────────────────
        # Yields (job_id, job_name, cleaned_page) tuples, streaming
        # pages from disk one at a time so only one page is in memory.
        async def _stream_pages():
            for jid, jname, on_disk, fpath in job_meta:
                if on_disk:
                    from app.utils.job_results_store import load_paginated_job_results_from_disk

                    offset = 0
                    while True:
                        page, total = await run_in_threadpool(
                            load_paginated_job_results_from_disk,
                            jid,
                            limit=_PAGINATION_CHUNK_SIZE,
                            offset=offset,
                            file_path=fpath,
                        )
                        if not page:
                            break
                        yield (jid, jname, _strip_system_fields(page))
                        offset += _PAGINATION_CHUNK_SIZE
                        if offset >= total:
                            break
                else:
                    job = jobs_store.get(jid)
                    if job and job.results:
                        # Cap in-memory results to prevent OOM
                        raw = list(job.results)[:10000]
                        yield (jid, jname, _strip_system_fields(raw))

        # ── Route to format-specific handler ────────────────────────────
        ts = datetime.datetime.now(datetime.UTC).strftime("%Y%m%d_%H%M%S")

        if fmt == "csv":
            return StreamingResponse(
                _batch_csv_stream(_stream_pages(), fieldnames, body.flatten, ts),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="batch_export_{ts}.csv"'},
            )
        if fmt == "json":
            return StreamingResponse(
                _batch_json_stream(_stream_pages(), fieldnames, body.flatten, ts),
                media_type="application/json",
                headers={"Content-Disposition": f'attachment; filename="batch_export_{ts}.json"'},
            )
        # fmt == "xlsx"  # noqa: ERA001, RUF100
        # XLSX requires full workbook in memory (openpyxl limitation).
        # Collect pages into per-job lists with a per-job cap.
        per_job_results: list[tuple[str, str, list[dict[str, Any]]]] = []
        current_jid: str | None = None
        current_rows: list[dict[str, Any]] = []
        current_name: str = ""
        async for jid, jname, page in _stream_pages():
            if jid != current_jid:
                if current_jid is not None:
                    per_job_results.append((current_jid, current_name, current_rows))
                current_jid = jid
                current_name = jname
                current_rows = []
            current_rows.extend(page)
            # Cap per-job to prevent OOM
            if len(current_rows) > 10000:
                current_rows = current_rows[:10000]
        if current_jid is not None:
            per_job_results.append((current_jid, current_name, current_rows))
        return await run_in_threadpool(_batch_xlsx, per_job_results, fieldnames, body.flatten)

    def _batch_xlsx(
        per_job_results: list[tuple[str, str, list[dict[str, Any]]]],
        fieldnames: list[str],
        flatten: bool,
    ) -> Response:
        """Generate a batch Excel response.

        When *flatten* is True, all rows go into a single "Combined"
        sheet with a ``_source_job`` column. When False, each job gets
        its own sheet named after the job (truncated to 31 chars).
        """
        wb = Workbook(write_only=True)

        def _row_values(row: dict[str, Any], fnames: list[str]) -> list[Any]:
            vals: list[Any] = []
            for f in fnames:
                v = row.get(f)
                if isinstance(v, list):
                    vals.append(_safe_cell(", ".join(str(i) for i in v if i is not None)))
                else:
                    vals.append(_safe_cell(v))
            return vals

        if flatten:
            all_fnames = list(fieldnames)
            all_fnames.append(_SOURCE_JOB_FIELD)
            # Use the same collision-avoidance logic as the per-job
            # branch below: a user-named job called exactly "Combined"
            # would otherwise raise ``InvalidWorksheetTitle`` from
            # openpyxl because the explicit title is duplicated with
            # itself across the same export call. Pre-register the
            # title in a one-element used-set so the loop is
            # unambiguous.
            used_flatten_names: set[str] = set()
            base = "Combined"
            sheet_name = base
            suffix = 2
            while sheet_name in used_flatten_names:
                candidate = f"{base[: 31 - 4]} (2)" if suffix == 2 else f"{base[: 31 - 4]} ({suffix})"
                sheet_name = candidate[:31]
                suffix += 1
                if suffix > 999:
                    sheet_name = f"{base[: 31 - 4]}_x"
                    break
            used_flatten_names.add(sheet_name)
            ws = wb.create_sheet(title=sheet_name)
            ws.append(all_fnames)
            for _, job_name, rows in per_job_results:
                for row in rows:
                    vals = _row_values(row, fieldnames)
                    vals.append(job_name)
                    ws.append(vals)
        else:
            # Track used sheet names to avoid ``InvalidWorksheetTitle`` from
            # openpyxl when two jobs share the same 31-char prefix.
            used_sheet_names: set[str] = set()
            for _, job_name, rows in per_job_results:
                base = (job_name or "Sheet")[:31] or "Sheet"
                sheet_name = base
                suffix = 2
                while sheet_name in used_sheet_names or sheet_name == "Sheet":
                    candidate = f"{base[: 31 - 4]} (2)" if suffix == 2 and base != "Sheet" else f"{base[: 31 - 4]} ({suffix})"
                    sheet_name = candidate[:31]
                    suffix += 1
                    if suffix > 999:
                        # Defensive cap — at this point we have hundreds of
                        # jobs with the same 31-char prefix, which is itself
                        # a data-quality issue, but we should not loop forever.
                        sheet_name = f"{base[: 31 - 4]}_x"
                        break
                used_sheet_names.add(sheet_name)
                ws = wb.create_sheet(title=sheet_name)
                ws.append(fieldnames)
                for row in rows:
                    ws.append(_row_values(row, fieldnames))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        ts = datetime.datetime.now(datetime.UTC).strftime("%Y%m%d_%H%M%S")
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="batch_export_{ts}.xlsx"'},
        )

    # ── Streaming batch generators (E1: memory-efficient) ────────────

    async def _batch_csv_stream(
        pages: AsyncIterator[tuple[str, str, list[dict[str, Any]]]],
        fieldnames: list[str],
        flatten: bool,
        _ts: str,
    ):
        """Yield CSV chunks as a streaming generator.

        Only one page of results is in memory at a time. The header is
        written first (from the discovered fieldnames), then each page
        is written row-by-row.
        """
        output = io.StringIO()
        writer: csv.DictWriter | None = None

        if flatten:
            all_fieldnames = list(fieldnames)
            all_fieldnames.append(_SOURCE_JOB_FIELD)
            writer = csv.DictWriter(output, fieldnames=all_fieldnames)
            writer.writeheader()
            yield output.getvalue()
            output.truncate(0)
            output.seek(0)

            async for _jid, job_name, page in pages:
                for row in page:
                    flat = _flat_row(row, fieldnames)
                    flat[_SOURCE_JOB_FIELD] = job_name
                    writer.writerow(flat)
                # Yield periodically to keep memory bounded
                chunk = output.getvalue()
                if chunk:
                    yield chunk
                    output.truncate(0)
                    output.seek(0)
        else:
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            yield output.getvalue()
            output.truncate(0)
            output.seek(0)

            first_job = True
            async for _jid, job_name, page in pages:
                if not first_job:
                    sep: dict[str, str] = dict.fromkeys(fieldnames, "")
                    sep[fieldnames[0]] = f"--- {job_name} ---"
                    writer.writerow(sep)
                first_job = False
                for row in page:
                    writer.writerow(_flat_row(row, fieldnames))
                chunk = output.getvalue()
                if chunk:
                    yield chunk
                    output.truncate(0)
                    output.seek(0)

    async def _batch_json_stream(
        pages: AsyncIterator[tuple[str, str, list[dict[str, Any]]]],
        fieldnames: list[str],  # noqa: ARG001
        flatten: bool,
        _ts: str,
    ):
        """Yield JSON chunks as a streaming generator.

        For flatten mode: yields a JSON array opening, then each row as
        a JSON object, then the closing bracket.
        For grouped mode: yields a JSON object with an ``exports`` array.
        """
        if flatten:
            yield "[\n"
            first = True
            async for _jid, job_name, page in pages:
                for row in page:
                    tagged = dict(row)
                    tagged[_SOURCE_JOB_FIELD] = job_name
                    if not first:
                        yield ",\n"
                    yield json.dumps(tagged)
                    first = False
            yield "\n]\n"
        else:
            yield '{\n  "exports": [\n'
            first_job = True
            async for jid, job_name, page in pages:
                if not first_job:
                    yield ",\n"
                first_job = False
                # Stream each job's results as a JSON object
                yield f"    {json.dumps({'job_id': jid, 'job_name': job_name, 'results': page})}"
            yield "\n  ]\n}\n"

    return router

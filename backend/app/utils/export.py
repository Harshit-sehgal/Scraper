import io
import re
from openpyxl import Workbook
from fastapi.responses import Response

def safe_export_filename(name: str, extension: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", (name or "").strip()).strip("._-")
    stem = stem[:80] or "dataforge_export"
    ext = re.sub(r"[^A-Za-z0-9]+", "", extension or "") or "dat"
    return f"{stem}.{ext}"

def generate_csv_response(job_name: str, fieldnames: list, results: list) -> Response:
    import csv
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(results)
    
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{safe_export_filename(job_name, "csv")}"'}
    )

def generate_json_response(job_name: str, results: list) -> Response:
    import json
    return Response(
        content=json.dumps(results, indent=2),
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{safe_export_filename(job_name, "json")}"'}
    )

def generate_excel_response(job_name: str, job_results: list, schema_fields: list) -> Response:
    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    fieldnames = [f.name for f in schema_fields] if schema_fields else (list(job_results[0].keys()) if job_results else [])

    # Write headers
    for col_num, header in enumerate(fieldnames, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num, row in enumerate(job_results, 2):
        for col_num, field in enumerate(fieldnames, 1):
            value = row.get(field)
            if isinstance(value, list):
                value = ", ".join(str(i) for i in value if i is not None)
            ws.cell(row=row_num, column=col_num, value=value)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_export_filename(job_name, "xlsx")}"'}
    )

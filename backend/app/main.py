"""
FastAPI Main Server — DataForge General-Purpose Web Scraper API.
"""

import asyncio
import csv
import io
import json
import datetime
import os
import re
import time
from statistics import mean
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from fastapi.staticfiles import StaticFiles
from pathlib import Path
from openpyxl import Workbook

from app.models import Job, JobCreate, JobStatus, ScrapeMode, DiscoveryRequest, SchemaSuggestionRequest, FieldType
from app.scraper import scrape_url, suggest_schema_from_intent, ai_clean_and_align_records
from app.filters import process_results, apply_location_radius
from app.discovery import discover_urls, infer_source_metadata
from app.state_store import load_state, save_state, get_state_file_path

app = FastAPI(
    title="DataForge — General-Purpose Web Scraper",
    description="AI-powered scraper that extracts structured data from any website",
    version="2.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def _env_int(name: str, default: int, minimum: int | None = None, maximum: int | None = None) -> int:
    raw = (os.getenv(name) or "").strip()
    if not raw:
        value = default
    else:
        try:
            value = int(raw)
        except Exception:
            value = default

    if minimum is not None:
        value = max(minimum, value)
    if maximum is not None:
        value = min(maximum, value)
    return value


# Runtime safety rails to avoid indefinitely-running jobs.
MAX_DISCOVERY_URLS = _env_int("DATAFORGE_MAX_DISCOVERY_URLS", 20, minimum=1, maximum=100)
PER_URL_SCRAPE_TIMEOUT_SECONDS = _env_int("DATAFORGE_PER_URL_TIMEOUT_SECONDS", 120, minimum=10, maximum=900)
MAX_JOB_RUNTIME_SECONDS = _env_int("DATAFORGE_MAX_JOB_RUNTIME_SECONDS", 1800, minimum=60, maximum=14400)
AI_STRUCTURING_TIMEOUT_SECONDS = _env_int("DATAFORGE_AI_STRUCTURING_TIMEOUT_SECONDS", 240, minimum=15, maximum=1800)
INSIGHT_TIMEOUT_SECONDS = _env_int("DATAFORGE_INSIGHT_TIMEOUT_SECONDS", 25, minimum=5, maximum=300)
MAX_JOB_HISTORY = _env_int("DATAFORGE_MAX_JOB_HISTORY", 300, minimum=25, maximum=5000)
MAX_RECYCLE_BIN_HISTORY = _env_int("DATAFORGE_MAX_RECYCLE_BIN_HISTORY", 300, minimum=25, maximum=5000)

# Durable job store (loaded at startup, persisted on mutations).
jobs_store, recycle_bin_store = load_state()
if jobs_store or recycle_bin_store:
    save_state(jobs_store=jobs_store, recycle_bin_store=recycle_bin_store)

# Serve Frontend
FRONTEND_DIR = Path(__file__).parent.parent.parent / "frontend"
if FRONTEND_DIR.exists():
    app.mount("/app", StaticFiles(directory=str(FRONTEND_DIR), html=True), name="frontend")


def _schedule_background_task(coro):
    return asyncio.create_task(coro)


@app.get("/")
async def root():
    return {"message": "DataForge API v2", "docs": "/docs", "dashboard": "/app"}


@app.get("/api/system/status")
async def system_status():
    counts = {s.value: 0 for s in JobStatus}
    for job in jobs_store.values():
        status_key = str(job.status.value if isinstance(job.status, JobStatus) else job.status)
        if status_key not in counts:
            counts[status_key] = 0
        counts[status_key] += 1

    active = counts.get(JobStatus.PENDING.value, 0) + counts.get(JobStatus.DISCOVERING.value, 0) + counts.get(JobStatus.RUNNING.value, 0)

    return {
        "status": "online",
        "jobs": {
            "total": len(jobs_store),
            "active": active,
            "completed": counts.get(JobStatus.COMPLETED.value, 0),
            "failed": counts.get(JobStatus.FAILED.value, 0),
            "canceled": counts.get(JobStatus.CANCELED.value, 0),
        },
        "runtime_limits": {
            "max_discovery_urls": MAX_DISCOVERY_URLS,
            "per_url_timeout_seconds": PER_URL_SCRAPE_TIMEOUT_SECONDS,
            "max_job_runtime_seconds": MAX_JOB_RUNTIME_SECONDS,
            "ai_structuring_timeout_seconds": AI_STRUCTURING_TIMEOUT_SECONDS,
            "insight_timeout_seconds": INSIGHT_TIMEOUT_SECONDS,
            "max_job_history": MAX_JOB_HISTORY,
            "max_recycle_bin_history": MAX_RECYCLE_BIN_HISTORY,
        },
        "state_file": str(get_state_file_path()),
    }


# ─── Discovery API ───

@app.post("/api/discover")
async def discover(req: DiscoveryRequest):
    """Auto-discover best URLs to scrape for a topic."""
    results = await discover_urls(
        query=req.topic,
        domain=req.domain,
        num_results=req.num_results,
        location=req.location,
        data_fields=req.schema_field_names,
        origin_location=req.origin_location,
        max_distance_km=req.max_distance_km,
        source_policy=req.source_policy,
        max_per_domain=req.max_per_domain,
    )
    return {"urls": results}


@app.post("/api/schema/suggest")
async def suggest_schema(req: SchemaSuggestionRequest):
    """Infer topic + schema fields from plain-language user intent."""
    suggestion = await suggest_schema_from_intent(req.intent, max_fields=req.max_fields)
    return suggestion


# ─── Jobs CRUD ───

@app.get("/api/jobs")
async def list_jobs():
    ordered = sorted(jobs_store.values(), key=lambda j: j.created_at, reverse=True)
    return {"jobs": [job.model_dump() for job in ordered]}


@app.get("/api/jobs/{job_id}")
async def get_job(job_id: str):
    if job_id not in jobs_store:
        raise HTTPException(status_code=404, detail="Job not found")
    job = jobs_store[job_id]
    if _backfill_source_metadata_for_job(job):
        _persist_state()
    return job.model_dump()


@app.post("/api/jobs")
async def create_job(job_data: JobCreate):
    manual_urls = [u.strip() for u in job_data.urls if str(u or "").strip()]
    urls = manual_urls if job_data.mode == ScrapeMode.MANUAL else []

    job = Job(
        name=job_data.name,
        mode=job_data.mode,
        intent=job_data.intent,
        urls=urls,
        topic=job_data.topic,
        location=job_data.location,
        preferred_domain=job_data.preferred_domain,
        source_policy=job_data.source_policy,
        max_per_domain=job_data.max_per_domain,
        origin_location=job_data.origin_location,
        max_distance_km=job_data.max_distance_km,
        schema_fields=job_data.schema_fields,
        filters=job_data.filters,
        pagination=job_data.pagination,
        max_pages=job_data.max_pages,
        deduplicate=job_data.deduplicate,
        deduplicate_field=job_data.deduplicate_field,
        min_record_score=job_data.min_record_score,
    )
    jobs_store[job.id] = job
    _persist_state()
    _schedule_background_task(_run_job(job.id))
    return {"job_id": job.id, "status": job.status.value}


@app.post("/api/jobs/{job_id}/cancel")
async def cancel_job(job_id: str):
    if job_id not in jobs_store:
        raise HTTPException(status_code=404, detail="Job not found")

    job = jobs_store[job_id]
    if job.status in {JobStatus.COMPLETED, JobStatus.FAILED, JobStatus.CANCELED}:
        return {
            "job_id": job.id,
            "status": job.status.value,
            "cancel_requested": bool(job.cancel_requested),
            "message": "Job already in terminal state",
        }

    job.cancel_requested = True
    if job.status == JobStatus.PENDING:
        _mark_job_canceled(job, "Canceled before execution.")

    _persist_state()
    return {
        "job_id": job.id,
        "status": job.status.value,
        "cancel_requested": True,
        "message": "Cancellation requested",
    }


@app.post("/api/jobs/{job_id}/reclean")
async def reclean_job(job_id: str):
    """Re-run AI cleaning and schema alignment on existing job results without re-scraping URLs."""
    if job_id not in jobs_store:
        raise HTTPException(status_code=404, detail="Job not found")

    job = jobs_store[job_id]
    if job.status in {JobStatus.PENDING, JobStatus.DISCOVERING, JobStatus.RUNNING}:
        raise HTTPException(status_code=409, detail="Job is still running; wait for completion before re-cleaning")
    if not job.results:
        raise HTTPException(status_code=400, detail="No results to re-clean")
    if not job.schema_fields:
        raise HTTPException(status_code=400, detail="Job has no schema fields for re-cleaning")

    started = datetime.datetime.now().isoformat()
    before_records = len(job.results)
    working_rows = [dict(r) for r in job.results]
    reclean_warnings: list[str] = []

    job.status = JobStatus.RUNNING

    ai_report = {
        "applied": False,
        "input_records": len(working_rows),
        "output_records": len(working_rows),
        "total_chunks": 0,
        "ai_chunks": 0,
        "fallback_chunks": 0,
        "model_fallback_mode": False,
        "noise_rows_removed": 0,
        "capped_records": 0,
        "quality_filtered_after_ai": 0,
    }

    try:
        cleaned_rows, ai_report = await asyncio.wait_for(
            ai_clean_and_align_records(
                working_rows,
                job.schema_fields,
                min_record_score=job.min_record_score,
            ),
            timeout=AI_STRUCTURING_TIMEOUT_SECONDS,
        )
    except asyncio.TimeoutError:
        cleaned_rows = working_rows
        reclean_warnings.append(
            f"AI re-clean timed out after {AI_STRUCTURING_TIMEOUT_SECONDS}s; used deterministic post-processing."
        )
    except Exception as e:
        cleaned_rows = working_rows
        reclean_warnings.append("AI re-clean failed; used deterministic post-processing.")
        print(f"[Job {job_id}] Re-clean failed: {e}")

    filtered_results, total, filtered_count, type_integrity_report = process_results(
        cleaned_rows,
        job.schema_fields,
        job.filters,
    )

    if job.deduplicate and filtered_results:
        filtered_results = _deduplicate_results(
            records=filtered_results,
            schema_fields=job.schema_fields,
            deduplicate_field=job.deduplicate_field,
        )
        filtered_count = len(filtered_results)

    # Backfill manual-mode source metadata so users don't only see "unknown".
    for row in filtered_results:
        source_url = str(row.get("source_url") or "")
        source_type = str(row.get("source_type") or "unknown").strip().lower()
        if source_type == "unknown" and source_url:
            inferred = infer_source_metadata(url=source_url)
            row["source_type"] = str(inferred.get("source_type") or "unknown")
            row["source_trust_score"] = round(_safe_score(inferred.get("source_trust_score") or 0.4), 3)

    job.results = normalize_job_results(filtered_results, job.schema_fields)
    job.total_records = total
    job.filtered_records = filtered_count
    job.completed_at = datetime.datetime.now().isoformat()
    job.status = JobStatus.COMPLETED

    # Update timestamps on refreshed records.
    scraped_at = datetime.datetime.now().isoformat()
    for row in job.results:
        row["scraped_at"] = scraped_at

    prev_quality = dict(job.quality_report or {})
    existing_warnings = list(prev_quality.get("warnings") or [])
    radius_report = prev_quality.get("radius")
    if not isinstance(radius_report, dict):
        radius_report = {
            "applied": False,
            "reason": "not_configured",
            "origin": job.origin_location,
            "max_distance_km": job.max_distance_km,
        }

    ai_source_prediction = prev_quality.get("ai_source_prediction")
    if not isinstance(ai_source_prediction, dict):
        ai_source_prediction = {
            "sources_attempted": 0,
            "sources_with_ai_structuring": 0,
            "records_processed": 0,
            "records_ai_structured": 0,
        }

    source_breakdown = _compute_source_breakdown(filtered_results)
    quality = build_quality_report(
        raw_results=cleaned_rows,
        post_filter_count=len(filtered_results),
        post_radius_count=len(filtered_results),
        radius_report=radius_report,
        final_results=filtered_results,
        min_record_score=job.min_record_score,
        type_integrity_report=type_integrity_report,
        source_breakdown=source_breakdown,
        ai_source_prediction=ai_source_prediction,
        ai_structuring_report=ai_report,
        warnings=[*existing_warnings, *reclean_warnings],
    )

    quality["reclean"] = {
        "applied": True,
        "started_at": started,
        "completed_at": job.completed_at,
        "before_records": before_records,
        "after_records": len(job.results),
        "ai_structuring": ai_report,
        "warnings": reclean_warnings,
    }
    job.quality_report = quality
    _persist_state()

    return {
        "job_id": job.id,
        "status": job.status.value,
        "before_records": before_records,
        "after_records": len(job.results),
        "warnings": reclean_warnings,
    }


@app.delete("/api/jobs/{job_id}")
async def delete_job(job_id: str):
    if job_id not in jobs_store:
        raise HTTPException(status_code=404, detail="Job not found")
    recycle_bin_store[job_id] = jobs_store.pop(job_id)
    _persist_state()
    return {"message": "Job moved to recycle bin"}


@app.delete("/api/jobs/cleanup/terminal")
async def clear_terminal_jobs(keep_recent: int = Query(5, ge=0, le=5000)):
    terminal_statuses = {JobStatus.COMPLETED, JobStatus.FAILED, JobStatus.CANCELED}
    terminal = [
        (jid, job)
        for jid, job in jobs_store.items()
        if job.status in terminal_statuses
    ]

    if not terminal:
        return {
            "message": "No terminal jobs to clear",
            "cleared": 0,
            "kept_recent": keep_recent,
            "remaining": len(jobs_store),
        }

    terminal.sort(key=lambda item: item[1].created_at, reverse=True)
    keep_ids = {jid for jid, _ in terminal[:keep_recent]}

    removed = 0
    for jid, _ in terminal:
        if jid in keep_ids:
            continue
        del jobs_store[jid]
        removed += 1

    if removed:
        _persist_state()

    return {
        "message": f"Cleared {removed} terminal jobs",
        "cleared": removed,
        "kept_recent": keep_recent,
        "remaining": len(jobs_store),
    }

@app.get("/api/recycle_bin")
async def list_recycle_bin():
    ordered = sorted(recycle_bin_store.values(), key=lambda j: j.created_at, reverse=True)
    return {"jobs": [job.model_dump() for job in ordered]}

@app.post("/api/recycle_bin/{job_id}/restore")
async def restore_job(job_id: str):
    if job_id not in recycle_bin_store:
        raise HTTPException(status_code=404, detail="Job not in recycle bin")
    jobs_store[job_id] = recycle_bin_store.pop(job_id)
    _persist_state()
    return {"message": "Job restored"}

@app.delete("/api/recycle_bin/{job_id}")
async def hard_delete_job(job_id: str):
    if job_id not in recycle_bin_store:
        raise HTTPException(status_code=404, detail="Job not in recycle bin")
    del recycle_bin_store[job_id]
    _persist_state()
    return {"message": "Job permanently deleted"}


@app.delete("/api/recycle_bin")
async def clear_recycle_bin():
    cleared = len(recycle_bin_store)
    recycle_bin_store.clear()
    if cleared:
        _persist_state()
    return {"message": f"Cleared {cleared} recycle-bin jobs", "cleared": cleared}


# ─── Export ───

@app.get("/api/jobs/{job_id}/export/csv")
async def export_csv(job_id: str):
    if job_id not in jobs_store:
        raise HTTPException(status_code=404, detail="Job not found")
    job = jobs_store[job_id]
    if not job.results:
        raise HTTPException(status_code=400, detail="No results to export")

    output = io.StringIO()
    fieldnames = [f.name for f in job.schema_fields] if job.schema_fields else (list(job.results[0].keys()) if job.results else [])
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row in job.results:
        flat_row = {}
        for k in fieldnames:
            v = row.get(k)
            if isinstance(v, list):
                flat_row[k] = ", ".join(str(i) for i in v)
            else:
                flat_row[k] = v
        writer.writerow(flat_row)

    output.seek(0)
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{_safe_export_filename(job.name, "csv")}"'}
    )


@app.get("/api/jobs/{job_id}/export/json")
async def export_json(job_id: str):
    if job_id not in jobs_store:
        raise HTTPException(status_code=404, detail="Job not found")
    job = jobs_store[job_id]
    if not job.results:
        raise HTTPException(status_code=400, detail="No results to export")

    json_content = json.dumps(job.results, indent=2)
    return Response(
        content=json_content,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{_safe_export_filename(job.name, "json")}"'}
    )


@app.get("/api/jobs/{job_id}/export/excel")
async def export_excel(job_id: str):
    if job_id not in jobs_store:
        raise HTTPException(status_code=404, detail="Job not found")
    job = jobs_store[job_id]
    if not job.results:
        raise HTTPException(status_code=400, detail="No results to export")

    wb = Workbook()
    ws = wb.active
    ws.title = "Scraped Data"

    fieldnames = [f.name for f in job.schema_fields] if job.schema_fields else (list(job.results[0].keys()) if job.results else [])

    # Write headers
    for col_num, header in enumerate(fieldnames, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num, row in enumerate(job.results, 2):
        for col_num, field in enumerate(fieldnames, 1):
            value = row.get(field)
            if isinstance(value, list):
                value = ", ".join(str(i) for i in value if i is not None)
            ws.cell(row=row_num, column=col_num, value=value)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{_safe_export_filename(job.name, "xlsx")}"'}
    )


# ─── Background Job Runner ───

def _prune_history_stores():
    if len(jobs_store) > MAX_JOB_HISTORY:
        active_ids = {
            jid
            for jid, job in jobs_store.items()
            if job.status not in {JobStatus.COMPLETED, JobStatus.FAILED, JobStatus.CANCELED}
        }
        slots_for_terminal = max(0, MAX_JOB_HISTORY - len(active_ids))

        terminal_jobs = [
            (jid, job)
            for jid, job in jobs_store.items()
            if jid not in active_ids
        ]
        terminal_jobs.sort(key=lambda item: item[1].created_at, reverse=True)

        keep_ids = set(active_ids)
        keep_ids.update(jid for jid, _ in terminal_jobs[:slots_for_terminal])

        for jid in list(jobs_store.keys()):
            if jid not in keep_ids:
                del jobs_store[jid]

    if len(recycle_bin_store) > MAX_RECYCLE_BIN_HISTORY:
        recycle_items = sorted(
            recycle_bin_store.items(),
            key=lambda item: item[1].created_at,
            reverse=True,
        )
        keep_ids = {jid for jid, _ in recycle_items[:MAX_RECYCLE_BIN_HISTORY]}
        for jid in list(recycle_bin_store.keys()):
            if jid not in keep_ids:
                del recycle_bin_store[jid]

def _persist_state():
    _prune_history_stores()
    save_state(jobs_store=jobs_store, recycle_bin_store=recycle_bin_store)


def _mark_job_canceled(job: Job, reason: str = "Canceled by user"):
    job.status = JobStatus.CANCELED
    job.error = reason
    job.completed_at = datetime.datetime.now().isoformat()


def _clamp01(value: float) -> float:
    return max(0.0, min(1.0, value))


def _safe_export_filename(name: str, extension: str) -> str:
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", (name or "").strip()).strip("._-")
    stem = stem[:80] or "dataforge_export"
    ext = re.sub(r"[^A-Za-z0-9]+", "", extension or "") or "dat"
    return f"{stem}.{ext}"


def normalize_job_results(results: list[dict], schema_fields: list):
    """Force consistent schema order in each record and keep extra keys after standard fields."""
    normalized = []
    for record in results:
        ordered = {f.name: record.get(f.name) for f in schema_fields}
        # keep extras in deterministic sorted order to avoid randomness
        for key in sorted(record.keys()):
            if key not in ordered:
                ordered[key] = record[key]
        normalized.append(ordered)
    return normalized


def _normalized_dedup_text(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip().casefold()


def _safe_score(value) -> float:
    try:
        return float(value)
    except Exception:
        return 0.0


def _compute_source_breakdown(rows: list[dict]) -> dict:
    breakdown = {
        "official": 0,
        "directory": 0,
        "social": 0,
        "search_result": 0,
        "unknown": 0,
    }
    for row in rows:
        st = str((row or {}).get("source_type") or "unknown")
        breakdown[st if st in breakdown else "unknown"] += 1
    return breakdown


def _backfill_source_metadata_for_job(job: Job) -> bool:
    """Ensure existing rows have inferred source_type/source_trust_score when source_url is present."""
    if not job.results:
        return False

    changed = False
    for row in job.results:
        source_url = str(row.get("source_url") or "").strip()
        if not source_url:
            continue

        source_type = str(row.get("source_type") or "unknown").strip().lower()
        trust_score = row.get("source_trust_score")
        if source_type != "unknown" and trust_score is not None:
            continue

        inferred = infer_source_metadata(url=source_url)
        row["source_type"] = str(inferred.get("source_type") or "unknown")
        row["source_trust_score"] = round(_safe_score(inferred.get("source_trust_score") or 0.4), 3)
        changed = True

    if changed:
        q = dict(job.quality_report or {})
        q["source_breakdown"] = _compute_source_breakdown(job.results)
        job.quality_report = q

    return changed


def _deduplicate_results(records: list[dict], schema_fields: list, deduplicate_field: str = "") -> list[dict]:
    if not records:
        return records

    dedup_key = deduplicate_field
    if not dedup_key and records:
        dedup_key = list(records[0].keys())[0]

    seen = set()
    unique = []
    for r in records:
        dedup_value = _normalized_dedup_text(r.get(dedup_key, "")) if dedup_key else ""
        if not dedup_value:
            dedup_value = "|".join(
                _normalized_dedup_text(r.get(f.name)) for f in schema_fields
            )

        if dedup_value and dedup_value not in seen:
            seen.add(dedup_value)
            unique.append(r)
        elif not dedup_value:
            unique.append(r)

    return unique


def build_quality_report(
    raw_results: list[dict],
    post_filter_count: int,
    post_radius_count: int,
    radius_report: dict,
    final_results: list[dict],
    min_record_score: float,
    type_integrity_report: dict,
    source_breakdown: dict,
    ai_source_prediction: dict | None = None,
    ai_structuring_report: dict | None = None,
    warnings: list[str] | None = None,
) -> dict:
    scores = [_safe_score(r.get("record_score", 0.0)) for r in raw_results if isinstance(r, dict)]
    kept_scores = [_safe_score(r.get("record_score", 0.0)) for r in final_results if isinstance(r, dict)]
    avg_score = round(mean(scores), 3) if scores else 0.0
    avg_final_score = round(mean(kept_scores), 3) if kept_scores else 0.0

    source_trust_scores = [_safe_score(r.get("source_trust_score", 0.4)) for r in final_results if isinstance(r, dict)]
    avg_source_trust = round(mean(source_trust_scores), 3) if source_trust_scores else 0.4

    coverage_ratio = round((len(final_results) / len(raw_results)), 3) if raw_results else (1.0 if final_results else 0.0)
    mismatch_count = int((type_integrity_report or {}).get("total_type_mismatches") or 0)
    mismatch_ratio = mismatch_count / max(1, len(final_results))

    # Weighted blend of quality score, retention, source trust, and type integrity.
    overall_score = round(
        _clamp01(
            (avg_final_score * 0.55)
            + (coverage_ratio * 0.2)
            + (avg_source_trust * 0.15)
            + ((1.0 - _clamp01(mismatch_ratio)) * 0.1)
        ),
        3,
    )

    if not final_results:
        overall_score = 0.0

    source_ai = dict(ai_source_prediction or {})
    processed = int(source_ai.get("records_processed") or 0)
    structured = int(source_ai.get("records_ai_structured") or 0)
    source_ai["ai_row_rate"] = round((structured / processed), 3) if processed else 0.0

    return {
        "raw_records": len(raw_results),
        "post_filter_records": post_filter_count,
        "post_radius_records": post_radius_count,
        "final_records": len(final_results),
        "overall_score": overall_score,
        "quality_threshold": min_record_score,
        "avg_record_score": avg_score,
        "avg_final_record_score": avg_final_score,
        "coverage_ratio": coverage_ratio,
        "avg_source_trust_score": avg_source_trust,
        "records_below_threshold": sum(1 for s in scores if s < min_record_score),
        "type_integrity": type_integrity_report,
        "source_breakdown": source_breakdown,
        "ai_source_prediction": source_ai,
        "ai_structuring": ai_structuring_report or {},
        "warnings": warnings or [],
        "radius": radius_report,
    }


async def _run_job(job_id: str):
    job = jobs_store.get(job_id)
    if not job:
        return

    all_raw_results = []
    warnings: list[str] = []
    ai_source_prediction = {
        "sources_attempted": 0,
        "sources_with_ai_structuring": 0,
        "records_processed": 0,
        "records_ai_structured": 0,
    }
    ai_structuring_report = {
        "applied": False,
        "input_records": 0,
        "output_records": 0,
        "total_chunks": 0,
        "ai_chunks": 0,
        "fallback_chunks": 0,
        "capped_records": 0,
        "quality_filtered_after_ai": 0,
    }
    started_at = time.monotonic()
    if not job.started_at:
        job.started_at = datetime.datetime.now().isoformat()

    if job.cancel_requested:
        _mark_job_canceled(job, "Canceled before execution.")
        _persist_state()
        return

    try:
        # Auto-discovery mode
        if job.mode == ScrapeMode.AUTO:
            job.status = JobStatus.DISCOVERING
            _persist_state()
            print(f"[Job {job_id}] Auto-discovering URLs for: {job.topic}")

            # In auto mode, reuse max_pages as discovery count and cap to runtime-safe limits.
            discovery_limit = int(job.max_pages or 10)
            discovery_limit = max(1, min(discovery_limit, MAX_DISCOVERY_URLS))
            
            discovered = await discover_urls(
                query=job.topic,
                domain=job.preferred_domain,
                num_results=discovery_limit,
                location=job.location,
                data_fields=[f.name for f in job.schema_fields],
                origin_location=job.origin_location,
                max_distance_km=job.max_distance_km,
                source_policy=job.source_policy,
                max_per_domain=job.max_per_domain,
            )
            job.discovered_urls = discovered
            job.urls = [d["url"] for d in discovered if "url" in d]

            if not job.urls:
                if job.cancel_requested:
                    _mark_job_canceled(job)
                else:
                    job.status = JobStatus.FAILED
                    job.error = "Could not discover any URLs for this topic"
                    job.completed_at = datetime.datetime.now().isoformat()
                _persist_state()
                return

            print(f"[Job {job_id}] Discovered {len(job.urls)} URLs")

            if job.cancel_requested:
                _mark_job_canceled(job)
                _persist_state()
                return

        job.status = JobStatus.RUNNING
        _persist_state()

        for idx, url in enumerate(job.urls, start=1):
            if job.cancel_requested:
                _mark_job_canceled(job)
                _persist_state()
                return

            elapsed = time.monotonic() - started_at
            if elapsed > MAX_JOB_RUNTIME_SECONDS:
                warnings.append(
                    f"Job runtime limit reached at {int(elapsed)}s; partial results returned."
                )
                print(f"[Job {job_id}] Runtime limit reached after {int(elapsed)}s")
                break

            try:
                results = await asyncio.wait_for(
                    scrape_url(url, job.schema_fields, min_record_score=job.min_record_score, user_intent=job.intent),
                    timeout=PER_URL_SCRAPE_TIMEOUT_SECONDS,
                )
                ai_source_prediction["sources_attempted"] += 1
                ai_structured_rows_for_source = 0
                for record in results:
                    if record.pop("_ai_source_structured", False):
                        ai_structured_rows_for_source += 1
                    record["source_url"] = url
                    source_type = "unknown"
                    source_trust_score = 0.4

                    if job.discovered_urls:
                        matched = next((d for d in job.discovered_urls if d.get("url") == url), None)
                        if matched:
                            source_type = str(matched.get("source_type") or "unknown")
                            source_trust_score = _safe_score(matched.get("source_trust_score") or 0.4)
                        else:
                            inferred = infer_source_metadata(url=url)
                            source_type = str(inferred.get("source_type") or "unknown")
                            source_trust_score = _safe_score(inferred.get("source_trust_score") or 0.4)
                    else:
                        inferred = infer_source_metadata(url=url)
                        source_type = str(inferred.get("source_type") or "unknown")
                        source_trust_score = _safe_score(inferred.get("source_trust_score") or 0.4)

                    record["source_type"] = source_type
                    record["source_trust_score"] = round(source_trust_score, 3)

                ai_source_prediction["records_processed"] += len(results)
                ai_source_prediction["records_ai_structured"] += ai_structured_rows_for_source
                if ai_structured_rows_for_source > 0:
                    ai_source_prediction["sources_with_ai_structuring"] += 1

                all_raw_results.extend(results)
                _persist_state()
            except asyncio.TimeoutError:
                msg = f"URL timeout skipped ({idx}/{len(job.urls)}): {url}"
                warnings.append(msg)
                print(f"[Job {job_id}] {msg}")
            except Exception as e:
                msg = f"URL scrape failed ({idx}/{len(job.urls)}): {url}"
                warnings.append(f"{msg} ({type(e).__name__})")
                print(f"[Job {job_id}] {msg}: {e}")
                continue

        run_global_ai_structuring = (
            bool(all_raw_results)
            and bool(job.schema_fields)
            and ai_source_prediction["sources_attempted"] == 0
        )

        if job.cancel_requested:
            _mark_job_canceled(job)
            _persist_state()
            return

        if run_global_ai_structuring:
            print(f"[Job {job_id}] AI structuring {len(all_raw_results)} scraped rows...")
            try:
                all_raw_results, ai_structuring_report = await asyncio.wait_for(
                    ai_clean_and_align_records(
                        all_raw_results,
                        job.schema_fields,
                        min_record_score=job.min_record_score,
                    ),
                    timeout=AI_STRUCTURING_TIMEOUT_SECONDS,
                )
                if ai_structuring_report.get("capped_records", 0) > 0:
                    warnings.append(
                        "AI structuring processed a capped subset of rows; "
                        "remaining rows used deterministic cleaning."
                    )
                if ai_structuring_report.get("model_fallback_mode"):
                    warnings.append(
                        "AI structuring switched to deterministic fallback after repeated model timeouts/errors."
                    )
            except asyncio.TimeoutError:
                warnings.append(
                    f"AI structuring timed out after {AI_STRUCTURING_TIMEOUT_SECONDS}s; "
                    "continuing with deterministic processing."
                )
                print(f"[Job {job_id}] AI structuring timed out")
            except Exception as struct_err:
                warnings.append("AI structuring failed; continuing with deterministic processing.")
                print(f"[Job {job_id}] AI structuring failed: {struct_err}")
        elif all_raw_results and job.schema_fields:
            ai_structuring_report = {
                "applied": False,
                "reason": "skipped_global_ai_source_level_applied",
                "input_records": len(all_raw_results),
                "output_records": len(all_raw_results),
                "total_chunks": 0,
                "ai_chunks": 0,
                "fallback_chunks": 0,
                "model_fallback_mode": False,
                "capped_records": 0,
                "quality_filtered_after_ai": 0,
            }

        # Post-process
        filtered_results, total, filtered_count, type_integrity_report = process_results(
            all_raw_results, job.schema_fields, job.filters
        )
        post_filter_count = len(filtered_results)

        # Optional radius filtering against origin location
        location_field = next((f.name for f in job.schema_fields if f.field_type.value == "location"), "")
        radius_report = {
            "applied": False,
            "reason": "not_configured",
            "origin": job.origin_location,
            "max_distance_km": job.max_distance_km,
        }
        if job.origin_location and job.max_distance_km is not None:
            filtered_results, radius_report = apply_location_radius(
                records=filtered_results,
                schema_fields=job.schema_fields,
                origin_address=job.origin_location,
                max_distance_km=job.max_distance_km,
                preferred_location_field=location_field,
            )
            filtered_count = len(filtered_results)
        post_radius_count = len(filtered_results)

        # Deduplication
        if job.deduplicate and filtered_results:
            filtered_results = _deduplicate_results(
                records=filtered_results,
                schema_fields=job.schema_fields,
                deduplicate_field=job.deduplicate_field,
            )
            filtered_count = len(filtered_results)

        source_breakdown = _compute_source_breakdown(filtered_results)

        has_contact_fields = any(
            field.field_type in {FieldType.EMAIL, FieldType.PHONE}
            for field in job.schema_fields
        )
        if has_contact_fields and ai_source_prediction["sources_attempted"] > 0:
            if ai_source_prediction["records_ai_structured"] == 0:
                if (os.getenv("GROQ_API_KEY") or "").strip():
                    warnings.append(
                        "AI source structuring covered 0% rows in this run; provider timeouts/rate limits may reduce phone/email extraction."
                    )
                else:
                    warnings.append(
                        "AI source structuring covered 0% rows in this run; set GROQ_API_KEY to improve phone/email extraction reliability."
                    )

        job.quality_report = build_quality_report(
            raw_results=all_raw_results,
            post_filter_count=post_filter_count,
            post_radius_count=post_radius_count,
            radius_report=radius_report,
            final_results=filtered_results,
            min_record_score=job.min_record_score,
            type_integrity_report=type_integrity_report,
            source_breakdown=source_breakdown,
            ai_source_prediction=ai_source_prediction,
            ai_structuring_report=ai_structuring_report,
            warnings=warnings,
        )

        job.results = normalize_job_results(filtered_results, job.schema_fields)
        job.total_records = total
        job.filtered_records = filtered_count
        
        # Add scraped_at timestamp to each record
        scraped_at = datetime.datetime.now().isoformat()
        for record in job.results:
            record["scraped_at"] = scraped_at
        
        # AI Insight Phase
        if job.results:
            if job.cancel_requested:
                _mark_job_canceled(job)
                _persist_state()
                return

            job.status = JobStatus.RUNNING
            print(f"[Job {job_id}] Generating AI insights over {len(job.results)} records...")
            try:
                from app.scraper import generate_data_insight
                analysis_text = await asyncio.wait_for(
                    generate_data_insight(job.results),
                    timeout=INSIGHT_TIMEOUT_SECONDS,
                )
                job.analysis = analysis_text
            except asyncio.TimeoutError:
                print(
                    f"[Job {job_id}] AI insight timed out after "
                    f"{INSIGHT_TIMEOUT_SECONDS}s; continuing without insight."
                )
                job.analysis = "Insight generation timed out."
            except Exception as ai_e:
                print(f"[Job {job_id}] AI insight generation failed: {ai_e}")
                
        job.status = JobStatus.COMPLETED
        job.cancel_requested = False
        job.completed_at = datetime.datetime.now().isoformat()
        _persist_state()

        print(f"[Job {job_id}] Completed: {total} total, {filtered_count} after filtering")

    except Exception as e:
        if job.cancel_requested:
            _mark_job_canceled(job)
            print(f"[Job {job_id}] Canceled")
        else:
            job.status = JobStatus.FAILED
            job.error = str(e)
            job.completed_at = datetime.datetime.now().isoformat()
            print(f"[Job {job_id}] Failed: {e}")
        _persist_state()

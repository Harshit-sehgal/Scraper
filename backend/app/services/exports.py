"""Pure export formatting service.

This module owns the byte/stream production for CSV, JSON, Excel, and
batch exports. It is intentionally free of FastAPI / middleware /
auth dependencies so it can be unit-tested in isolation and so the
HTTP layer in ``app.routers.exports`` stays a thin adapter.

Public surface
--------------

CSV
~~~
- :func:`build_csv_response` -- in-memory CSV for a single job
- :func:`stream_csv_from_disk` -- async iterator streaming CSV from
  paginated disk reads
- :func:`flat_row` -- flatten a single result row for CSV writing
- :func:`safe_cell` -- escape formula-injection prefixes
- :func:`user_fieldnames` / :func:`strip_system_fields` -- helpers
  for schema-less exports

JSON
~~~~
- :func:`build_json_response` -- in-memory JSON for a single job
- :func:`stream_json_from_disk` -- async iterator streaming JSON from
  paginated disk reads

Excel
~~~~~
- :func:`build_excel_bytes` -- XLSX bytes for a single job (in-memory
  results or paginated disk reads)
- :func:`batch_xlsx` -- XLSX bytes for a multi-job batch export with
  optional flatten mode and a Summary sheet

Batch
~~~~~
- :func:`build_batch_manifest` -- build the per-job manifest entry list
- :func:`discover_fieldnames_union` -- union of fieldnames across jobs
- :func:`make_unique_sheet_name` -- 31-char-capped, collision-safe
- :func:`batch_csv_stream` -- async iterator for batch CSV
- :func:`batch_json_stream` -- async iterator for batch JSON
"""

from __future__ import annotations

import csv
import datetime
import io
import json
from collections.abc import AsyncIterator, Callable, Iterable, Iterator
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook

if TYPE_CHECKING:
    from app.models import Job


SOURCE_JOB_FIELD = "_source_job"
"""Field name injected into each result row to identify the source job."""

PAGINATION_CHUNK_SIZE = 500
"""Default page size for paginated disk reads."""

# Spreadsheet formula-injection prefixes. Values that begin with one of
# these strings get a leading single quote prepended on export so
# Excel / Sheets / LibreOffice treat them as plain text.
DANGEROUS_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

# In-memory cap for batch exports from a single in-memory job, to
# bound peak memory when callers have huge in-memory result lists.
INMEMORY_BATCH_CAP = 10000


# ════════════════════════════════════════════════════════════════════
# Cell-safety helpers
# ════════════════════════════════════════════════════════════════════


def safe_cell(value: Any) -> Any:
    """Escape spreadsheet formula-injection prefixes in *value*.

    Strings that start with ``=``, ``+``, ``-``, ``@``, tab, or CR
    get a single quote prepended. Non-strings are returned unchanged.
    """
    if isinstance(value, str) and value.startswith(DANGEROUS_PREFIXES):
        return "'" + value
    return value


def strip_system_fields(records: list[dict]) -> list[dict]:
    """Return a copy of *records* with all ``_``-prefixed keys removed."""
    return [{k: v for k, v in r.items() if not k.startswith("_")} for r in records]


def user_fieldnames(results_list: list[dict]) -> list[str]:
    """Return field names from the first record, filtering ``_``-prefixed internals."""
    if not results_list:
        return []
    return [k for k in results_list[0] if not k.startswith("_")]


def flat_row(row: dict[str, Any], fieldnames: list[str]) -> dict[str, Any]:
    """Flatten list values in *row* to comma-separated strings and escape injection."""
    flat: dict[str, Any] = {}
    for k in fieldnames:
        v = row.get(k)
        if isinstance(v, list):
            flat[k] = safe_cell(", ".join(str(i) for i in v))
        else:
            flat[k] = safe_cell(v)
    return flat


def resolve_fieldnames(job: Job, results_list: list[dict]) -> list[str]:
    """Return field names from a job's schema, or infer from the first record."""
    schema_fields = getattr(job, "schema_fields", None) or []
    if schema_fields:
        return [f.name for f in schema_fields]
    return user_fieldnames(results_list)


# ════════════════════════════════════════════════════════════════════
# CSV (single job)
# ════════════════════════════════════════════════════════════════════


def build_csv_bytes(job: Job) -> bytes:
    """Build CSV bytes for a job that has in-memory results."""
    if not job.results:
        msg = "No results to export"
        raise ValueError(msg)
    fieldnames = resolve_fieldnames(job, list(job.results))
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row in job.results:
        writer.writerow(flat_row(row, fieldnames))
    return output.getvalue().encode("utf-8")


async def stream_csv_from_disk(
    job: Job,
    page_loader: Callable[..., tuple[list[dict], int]],
) -> AsyncIterator[bytes]:
    """Stream CSV bytes for a job that has results on disk.

    *page_loader* is a callable with signature
    ``(job_id, limit, offset, file_path) -> (page, total)`` — the
    router passes ``load_paginated_job_results_from_disk`` directly.
    """
    first_page, total = await _call_async(
        page_loader,
        job.id,
        limit=PAGINATION_CHUNK_SIZE,
        offset=0,
        file_path=getattr(job, "results_file_path", None),
    )
    if not first_page:
        msg = "No results to export"
        raise ValueError(msg)

    fieldnames = resolve_fieldnames(job, first_page)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    yield output.getvalue().encode("utf-8")
    output.seek(0)
    output.truncate()

    for row in first_page:
        writer.writerow(flat_row(row, fieldnames))
    yield output.getvalue().encode("utf-8")
    output.seek(0)
    output.truncate()

    offset = PAGINATION_CHUNK_SIZE
    while offset < total:
        page, _ = await _call_async(
            page_loader,
            job.id,
            limit=PAGINATION_CHUNK_SIZE,
            offset=offset,
            file_path=getattr(job, "results_file_path", None),
        )
        if not page:
            break
        for row in page:
            writer.writerow(flat_row(row, fieldnames))
        yield output.getvalue().encode("utf-8")
        output.seek(0)
        output.truncate()
        offset += PAGINATION_CHUNK_SIZE


# ════════════════════════════════════════════════════════════════════
# JSON (single job)
# ════════════════════════════════════════════════════════════════════


def build_json_bytes(job: Job) -> bytes:
    """Build pretty-printed JSON bytes for a job that has in-memory results."""
    if not job.results:
        msg = "No results to export"
        raise ValueError(msg)
    cleaned = strip_system_fields(list(job.results))
    return json.dumps(cleaned, indent=2).encode("utf-8")


async def stream_json_from_disk(
    job: Job,
    page_loader: Callable[..., tuple[list[dict], int]],
) -> AsyncIterator[bytes]:
    """Stream pretty-printed JSON bytes for a job that has results on disk."""
    first_page, total = await _call_async(
        page_loader,
        job.id,
        limit=PAGINATION_CHUNK_SIZE,
        offset=0,
        file_path=getattr(job, "results_file_path", None),
    )
    if not first_page:
        msg = "No results to export"
        raise ValueError(msg)

    yield b"[\n"
    for i, row in enumerate(first_page):
        cleaned = strip_system_fields([row])[0]
        prefix = "  " if i == 0 else ",  "
        yield (prefix + json.dumps(cleaned, indent=2).replace("\n", "\n  ") + "\n").encode("utf-8")

    offset = PAGINATION_CHUNK_SIZE
    idx = len(first_page)
    while offset < total:
        page, _ = await _call_async(
            page_loader,
            job.id,
            limit=PAGINATION_CHUNK_SIZE,
            offset=offset,
            file_path=getattr(job, "results_file_path", None),
        )
        if not page:
            break
        for row in page:
            cleaned = strip_system_fields([row])[0]
            prefix = ",  " if idx > 0 else "  "
            yield (prefix + json.dumps(cleaned, indent=2).replace("\n", "\n  ") + "\n").encode("utf-8")
            idx += 1
        offset += PAGINATION_CHUNK_SIZE
    yield b"]\n"


# ════════════════════════════════════════════════════════════════════
# Excel (single job)
# ════════════════════════════════════════════════════════════════════


def _excel_value(row: dict[str, Any], field: str) -> Any:
    """Return a single value formatted for an Excel cell, escaping formula injection."""
    v = row.get(field)
    if isinstance(v, list):
        return safe_cell(", ".join(str(i) for i in v if i is not None))
    return safe_cell(v)


def build_excel_bytes(
    job: Job,
    page_loader: Callable[..., tuple[list[dict], int]] | None = None,
) -> bytes:
    """Build XLSX bytes for a single job.

    When *page_loader* is provided the job is treated as
    ``results_on_disk=True`` and the loader is used to page through
    results in chunks of :data:`PAGINATION_CHUNK_SIZE`. When omitted,
    the in-memory ``job.results`` list is used.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Scraped Data")

    if page_loader is not None:
        first_page, total = page_loader(
            job.id,
            limit=PAGINATION_CHUNK_SIZE,
            offset=0,
            file_path=getattr(job, "results_file_path", None),
        )
        if not first_page:
            msg = "No results to export"
            raise ValueError(msg)
        fieldnames = resolve_fieldnames(job, first_page)
        ws.append(fieldnames)
        for row in first_page:
            ws.append([_excel_value(row, f) for f in fieldnames])
        offset = PAGINATION_CHUNK_SIZE
        while offset < total:
            page, _ = page_loader(
                job.id,
                limit=PAGINATION_CHUNK_SIZE,
                offset=offset,
                file_path=getattr(job, "results_file_path", None),
            )
            if not page:
                break
            for row in page:
                ws.append([_excel_value(row, f) for f in fieldnames])
            offset += PAGINATION_CHUNK_SIZE
    else:
        results_list = list(job.results) if job.results else []
        if not results_list:
            msg = "No results to export"
            raise ValueError(msg)
        fieldnames = resolve_fieldnames(job, results_list)
        ws.append(fieldnames)
        for row in results_list:
            ws.append([_excel_value(row, f) for f in fieldnames])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ════════════════════════════════════════════════════════════════════
# Sheet-name collision
# ════════════════════════════════════════════════════════════════════


def make_unique_sheet_name(base: str, used: set[str], *, max_len: int = 31) -> str:
    """Return a sheet name based on *base* that is not in *used* and fits *max_len*.

    Excel's hard cap on sheet-name length is 31 characters. When
    *base* already appears in *used*, the function appends a numeric
    suffix (e.g. ``" (2)"``, ``" (3)"`` …) and re-checks. After 999
    suffix attempts it falls back to ``"<base>_x"``.
    """
    sheet_name = (base or "Sheet")[:max_len] or "Sheet"
    suffix = 2
    while sheet_name in used or sheet_name == "Sheet":
        # Match the legacy logic: when this is the very first
        # collision and the original base was not "Sheet", use a
        # (2) suffix; otherwise use a numeric suffix and shrink the
        # base so the suffix fits in 31 chars.
        if suffix == 2 and sheet_name != "Sheet":
            candidate = f"{sheet_name[: max_len - 4]} (2)"
        else:
            candidate = f"{sheet_name[: max_len - 4]} ({suffix})"
        sheet_name = candidate[:max_len]
        suffix += 1
        if suffix > 999:
            sheet_name = f"{sheet_name[: max_len - 2]}_x"
            break
    used.add(sheet_name)
    return sheet_name


# ════════════════════════════════════════════════════════════════════
# Batch manifest + fieldname union
# ════════════════════════════════════════════════════════════════════


async def build_batch_manifest(
    job_meta: list[tuple[str, str, bool, str | None]],
    page_loader: Callable[..., tuple[list[dict], int]] | None,
    get_inmemory_count: Callable[[str], int | None],
) -> list[dict[str, Any]]:
    """Build a per-job manifest list for a batch export.

    Each entry contains ``job_id``, ``job_name``, ``status``
    (``included`` / ``empty``), ``record_count``, and an optional
    ``truncated`` / ``original_count`` for in-memory jobs above
    :data:`INMEMORY_BATCH_CAP`.
    """
    manifest: list[dict[str, Any]] = []
    for jid, jname, on_disk, fpath in job_meta:
        entry: dict[str, Any] = {"job_id": jid, "job_name": jname, "status": "pending", "record_count": 0}
        if on_disk and page_loader is not None:
            first_page, total = await _call_async(page_loader, jid, limit=1, offset=0, file_path=fpath)
            if first_page:
                entry["status"] = "included"
                entry["record_count"] = total
            else:
                entry["status"] = "empty"
        else:
            count = get_inmemory_count(jid)
            if count is not None and count > 0:
                entry["status"] = "included"
                entry["record_count"] = count
                if count > INMEMORY_BATCH_CAP:
                    entry["truncated"] = True
                    entry["original_count"] = count
            else:
                entry["status"] = "empty"
        manifest.append(entry)
    return manifest


async def discover_fieldnames_union(
    job_meta: list[tuple[str, str, bool, str | None]],
    page_loader: Callable[..., tuple[list[dict], int]] | None,
    get_inmemory_sample: Callable[[str], list[dict]],
) -> tuple[list[str], bool]:
    """Return ``(union_of_fieldnames, has_any_data)`` across all batch jobs."""
    fieldnames: list[str] = []
    seen: set[str] = set()
    has_any_data = False

    # First pass: from the first page / sample of each job.
    for jid, _jname, on_disk, fpath in job_meta:
        if on_disk and page_loader is not None:
            first_page, _ = await _call_async(
                page_loader,
                jid,
                limit=PAGINATION_CHUNK_SIZE,
                offset=0,
                file_path=fpath,
            )
            if first_page:
                has_any_data = True
                for row in strip_system_fields(first_page):
                    for k in row:
                        if k not in seen:
                            seen.add(k)
                            fieldnames.append(k)
        else:
            sample = get_inmemory_sample(jid)
            if sample:
                has_any_data = True
                for row in strip_system_fields(sample):
                    for k in row:
                        if k not in seen:
                            seen.add(k)
                            fieldnames.append(k)

    # Second pass: subsequent pages of disk jobs.
    for jid, _jname, on_disk, fpath in job_meta:
        if not (on_disk and page_loader is not None):
            continue
        offset = PAGINATION_CHUNK_SIZE
        while True:
            page, total = await _call_async(
                page_loader,
                jid,
                limit=PAGINATION_CHUNK_SIZE,
                offset=offset,
                file_path=fpath,
            )
            if not page:
                break
            for row in strip_system_fields(page):
                for k in row:
                    if k not in seen:
                        seen.add(k)
                        fieldnames.append(k)
            offset += PAGINATION_CHUNK_SIZE
            if offset >= total:
                break

    return fieldnames, has_any_data


async def iter_batch_pages(
    job_meta: list[tuple[str, str, bool, str | None]],
    page_loader: Callable[..., tuple[list[dict], int]] | None,
    get_inmemory_results: Callable[[str], list[dict] | None],
) -> AsyncIterator[tuple[str, str, list[dict]]]:
    """Yield ``(job_id, job_name, cleaned_page)`` tuples for a batch export.

    For disk-backed jobs, yields each paginated page until the job
    is exhausted. For in-memory jobs, yields a single (capped) page
    containing up to :data:`INMEMORY_BATCH_CAP` rows. Jobs with no
    data are silently skipped.
    """
    for jid, jname, on_disk, fpath in job_meta:
        if on_disk and page_loader is not None:
            offset = 0
            while True:
                page, total = await _call_async(
                    page_loader,
                    jid,
                    limit=PAGINATION_CHUNK_SIZE,
                    offset=offset,
                    file_path=fpath,
                )
                if not page:
                    break
                yield (jid, jname, strip_system_fields(page))
                offset += PAGINATION_CHUNK_SIZE
                if offset >= total:
                    break
        else:
            results = get_inmemory_results(jid)
            if results:
                yield (jid, jname, strip_system_fields(list(results)[:INMEMORY_BATCH_CAP]))


# ════════════════════════════════════════════════════════════════════
# Batch format helpers
# ════════════════════════════════════════════════════════════════════


async def batch_csv_stream(
    pages: AsyncIterator[tuple[str, str, list[dict]]],
    fieldnames: list[str],
    flatten: bool,
) -> AsyncIterator[bytes]:
    """Yield CSV bytes for a multi-job batch export.

    When *flatten* is True a single ``_source_job`` column is added;
    when False, a ``--- <job_name> ---`` separator row is inserted
    between jobs.
    """
    output = io.StringIO()
    if flatten:
        all_fieldnames = [*list(fieldnames), SOURCE_JOB_FIELD]
        writer = csv.DictWriter(output, fieldnames=all_fieldnames)
        writer.writeheader()
        yield output.getvalue().encode("utf-8")
        output.seek(0)
        output.truncate()
        async for _jid, job_name, page in pages:
            for row in page:
                flat = flat_row(row, fieldnames)
                flat[SOURCE_JOB_FIELD] = job_name
                writer.writerow(flat)
            chunk = output.getvalue()
            if chunk:
                yield chunk.encode("utf-8")
                output.truncate(0)
                output.seek(0)
    else:
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        yield output.getvalue().encode("utf-8")
        output.seek(0)
        output.truncate()
        first_job = True
        async for _jid, job_name, page in pages:
            if not first_job:
                sep: dict[str, str] = dict.fromkeys(fieldnames, "")
                sep[fieldnames[0]] = f"--- {job_name} ---"
                writer.writerow(sep)
            first_job = False
            for row in page:
                writer.writerow(flat_row(row, fieldnames))
            chunk = output.getvalue()
            if chunk:
                yield chunk.encode("utf-8")
                output.truncate(0)
                output.seek(0)


async def batch_json_stream(
    pages: AsyncIterator[tuple[str, str, list[dict]]],
    flatten: bool,
    manifest: list[dict[str, Any]] | None = None,
) -> AsyncIterator[bytes]:
    """Yield JSON bytes for a multi-job batch export."""
    if flatten:
        yield b"[\n"
        first = True
        async for _jid, job_name, page in pages:
            for row in page:
                tagged = dict(row)
                tagged[SOURCE_JOB_FIELD] = job_name
                if not first:
                    yield b",\n"
                yield json.dumps(tagged).encode("utf-8")
                first = False
        yield b"\n]\n"
    else:
        yield b'{\n  "manifest": '
        yield json.dumps(manifest or []).encode("utf-8")
        yield b',\n  "exports": [\n'
        first_job = True
        async for jid, job_name, page in pages:
            if not first_job:
                yield b",\n"
            first_job = False
            yield (f"    {json.dumps({'job_id': jid, 'job_name': job_name, 'results': page})}".encode())
        yield b"\n  ]\n}\n"


def batch_xlsx(
    job_meta: list[tuple[str, str, bool, str | None]],
    fieldnames: list[str],
    flatten: bool,
    manifest: list[dict[str, Any]] | None = None,
    page_loader: Callable[..., tuple[list[dict], int]] | None = None,
    get_inmemory_results: Callable[[str], list[dict] | None] | None = None,
) -> bytes:
    """Build XLSX bytes for a multi-job batch export.

    *page_loader* and *get_inmemory_results* mirror the contract
    used elsewhere; the legacy router used an inline generator
    helper, this service version takes two callables for testability.
    """
    wb = Workbook(write_only=True)

    # Summary sheet (E2) with per-job status.
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.append(["Job ID", "Job Name", "Status", "Record Count", "Truncated"])
    for entry in manifest or []:
        ws_summary.append(
            [
                entry.get("job_id", ""),
                entry.get("job_name", ""),
                entry.get("status", ""),
                str(entry.get("record_count", 0)),
                str(entry.get("truncated", False)),
            ],
        )

    def _row_values(row: dict[str, Any], fnames: list[str]) -> list[Any]:
        return [_excel_value(row, f) for f in fnames]

    def _stream_pages_sync() -> Iterator[tuple[str, str, list[dict]]]:
        for jid, jname, on_disk, fpath in job_meta:
            rows_yielded = 0
            if on_disk and page_loader is not None:
                offset = 0
                while rows_yielded < INMEMORY_BATCH_CAP:
                    limit = min(PAGINATION_CHUNK_SIZE, INMEMORY_BATCH_CAP - rows_yielded)
                    page, total = page_loader(jid, limit=limit, offset=offset, file_path=fpath)
                    if not page:
                        break
                    yield jid, jname, strip_system_fields(page)
                    rows_yielded += len(page)
                    offset += len(page)
                    if offset >= total:
                        break
            else:
                if get_inmemory_results is None:
                    continue
                results = get_inmemory_results(jid)
                if results:
                    yield jid, jname, strip_system_fields(list(results)[:INMEMORY_BATCH_CAP])

    if flatten:
        all_fnames = [*list(fieldnames), SOURCE_JOB_FIELD]
        used_flatten_names: set[str] = set()
        # Pre-register "Combined" so the loop's make_unique_sheet_name
        # call is consistent with the inline logic that previously
        # lived in the router.
        base = "Combined"
        sheet_name = make_unique_sheet_name(base, used_flatten_names)
        ws = wb.create_sheet(title=sheet_name)
        ws.append(all_fnames)
        for _, job_name, page in _stream_pages_sync():
            for row in page:
                vals = _row_values(row, fieldnames)
                vals.append(job_name)
                ws.append(vals)
    else:
        used_sheet_names: set[str] = set()
        current_ws = None
        current_jid = None
        for jid, job_name, page in _stream_pages_sync():
            if jid != current_jid:
                current_jid = jid
                base = (job_name or "Sheet")[:31] or "Sheet"
                sheet_name = make_unique_sheet_name(base, used_sheet_names)
                current_ws = wb.create_sheet(title=sheet_name)
                current_ws.append(fieldnames)
            if current_ws is not None:
                for row in page:
                    current_ws.append(_row_values(row, fieldnames))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ════════════════════════════════════════════════════════════════════
# Small helpers
# ════════════════════════════════════════════════════════════════════


def batch_export_timestamp(now: datetime.datetime | None = None) -> str:
    """Return the canonical ``YYYYMMDD_HHMMSS`` stamp used in batch filenames."""
    return (now or datetime.datetime.now(datetime.UTC)).strftime("%Y%m%d_%H%M%S")


async def _call_async(func: Callable[..., Any], *args: Any, **kwargs: Any) -> Any:
    """Call *func* awaiting it; if it's a regular function, run it in the current loop.

    The router's *page_loader* is a regular ``def`` that uses
    ``run_in_threadpool``; here we accept either a coroutine function
    or a regular function and normalize the call. Async functions are
    awaited directly; sync functions are called and returned.
    """
    if hasattr(func, "__code__") and func.__code__.co_flags & 0x100:  # CO_COROUTINE
        return await func(*args, **kwargs)
    result = func(*args, **kwargs)
    if hasattr(result, "__await__"):
        return await result
    return result


def get_inmemory_results_map(jobs_store: dict[str, Any], job_ids: Iterable[str]) -> dict[str, list[dict] | None]:
    """Build a ``job_id -> results`` map for in-memory lookups (batch helpers)."""
    res: dict[str, list[dict] | None] = {}
    for jid in job_ids:
        job = jobs_store.get(jid)
        res[jid] = job.results if job is not None else None
    return res
